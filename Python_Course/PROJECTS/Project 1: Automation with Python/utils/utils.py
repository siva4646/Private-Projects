from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference

def process_workbook(file_name, row, max_row, col, max_col):
    wb = load_workbook(file_name)
    ws = wb.active
    for row in range(2, (max_row + 1)):
        cell = ws.cell(row, 3)
        print(cell.value)
        corrected_price = cell.value * 0.9
        corrected_price_cell = ws.cell(row, 4)
        corrected_price_cell.value = corrected_price


    values = Reference(ws
                       , min_row=row
                       , max_row=max_row
                       , min_col=col
                       , max_col=max_col
                       )

    chart = BarChart()
    chart.add_data(values)
    ws.add_chart(chart, 'e2')

    wb.save(file_name)